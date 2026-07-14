# export.py - Lead Export Module
# ============================================================
# VERSION : v1.0
# PURPOSE : Joins scored leads to property details, producing
#           the investor-ready ranked list. Writes two outputs:
#
#           1. Excel file (veridex_scored_leads.xlsx)
#              → Power BI data source + investor deliverable
#              → Contains a named Excel Table ("ScoredLeads")
#                that Power BI auto-detects on import
#              → Frozen headers, auto-width columns, sorted by rank
#
#           2. Materialized DB table (export_leads)
#              → Canonical scored dataset in SQLite
#              → Backup data source for Power BI via ODBC
#              → Available for ad-hoc SQL queries
#
#           Called by main.py via run_export(state) after predictor.
#
# POWER BI SETUP (one-time, after first pipeline run):
#   1. Open Power BI Desktop
#   2. Get Data → Excel Workbook
#   3. Navigate to veridex_scored_leads.xlsx in your project folder
#   4. Select the "ScoredLeads" table → Load
#   5. After each pipeline re-run, click Refresh in Power BI
#      to pull the updated data
#
# EXPORT STEPS (in order):
#   Step 1 → Load scored leads (state['scored_df'] or scored_leads table)
#   Step 2 → Load property details from leads table
#   Step 3 → Join on OBJECTID — single flat denormalized table
#   Step 4 → Add temperature band (HOT / WARM / COLD)
#   Step 5 → Reorder and cast columns for Power BI
#   Step 6 → Save to export_leads table in database
#   Step 7 → Write formatted Excel file with named table
#
# STATE CONTRACT:
#   Reads  → state['scored_df']   (DataFrame from predictor)
#   Reads  → state['run_id']      (current run ID)
#   Reads  → state['db_path']     (path to SQLite database)
#   Writes → state['export_df']   (final export DataFrame)
#   Errors → state['errors']      (appended on failure)
# ============================================================

import os
import sqlite3
import logging
import pandas as pd

import project_config as config
import database

logger = logging.getLogger(__name__)

# ── Export settings ───────────────────────────────────────────
EXCEL_FILENAME = "veridex_scored_leads.xlsx"
EXCEL_SHEET    = "Scored Leads"
TABLE_NAME     = "ScoredLeads"    # Power BI auto-detects named tables

# Columns to pull from the leads table for the export join
LEAD_DETAIL_COLUMNS = [
    "OBJECTID",
    "OWN_NAME",
    "OWN_ADDR1",
    "OWN_STATE_",
    "Parcel",
    "PHY_CITY",
    "PHY_ZIPCD",
    "JV",
    "LND_VAL",
    "ACT_YR_BLT",
    "TOT_LVG_AR",
    "SALE_PRC1",
    "SALE_YR1",
    "NO_BULDNG",
    "DOR_UC",
    "NBRHD_CD",
]

# Final column order in the export — rank and score first,
# property details in the middle, OBJECTID last (reference key).
# This order determines what Power BI shows by default.
EXPORT_COLUMNS = [
    "rank",
    "hot_lead_score",
    "temperature",
    "OWN_NAME",
    "Parcel",
    "OWN_STATE_",
    "PHY_CITY",
    "PHY_ZIPCD",
    "JV",
    "LND_VAL",
    "ACT_YR_BLT",
    "TOT_LVG_AR",
    "SALE_PRC1",
    "SALE_YR1",
    "NO_BULDNG",
    "DOR_UC",
    "NBRHD_CD",
    "OBJECTID",
]

# Temperature thresholds — presentation layer only.
# These do NOT affect scoring or model output.
TEMP_HOT_THRESHOLD  = 0.8
TEMP_WARM_THRESHOLD = 0.5


# ============================================================
# INDIVIDUAL EXPORT STEPS
# ============================================================

def _step1_load_scores(state: dict) -> pd.DataFrame:
    """
    Step 1 — Load scored leads from state or database.

    Prefers state['scored_df'] written by the predictor in the
    same run. Falls back to loading the most recent run from
    the scored_leads table if state is empty.
    """
    df = state.get("scored_df")

    if df is not None and len(df) > 0:
        logger.info(
            f"Step 1 | Load scores        | "
            f"{len(df):,} scored leads from state"
        )
        return df

    # Fallback — load from DB
    logger.info(
        "scored_df is empty. Loading from scored_leads table."
    )
    run_id = state.get("run_id")
    df = database.load_scored_leads(run_id)

    if df is None or len(df) == 0:
        raise ValueError(
            "No scored leads found. Run PREDICT mode first."
        )

    logger.info(
        f"Step 1 | Load scores        | "
        f"{len(df):,} scored leads from database"
    )
    return df


def _step2_load_lead_details(state: dict) -> pd.DataFrame:
    """
    Step 2 — Load property details from the leads table.

    Selects only the columns needed for the export join.
    The leads table contains the full raw dataset —
    we pull the investor-relevant subset.
    """
    conn = sqlite3.connect(state["db_path"])
    cols_sql = ", ".join(LEAD_DETAIL_COLUMNS)
    leads = pd.read_sql_query(
        f"SELECT {cols_sql} FROM leads", conn
    )
    conn.close()

    logger.info(
        f"Step 2 | Load lead details  | "
        f"{len(leads):,} records from leads table"
    )
    return leads


def _step3_join(scores: pd.DataFrame, leads: pd.DataFrame) -> pd.DataFrame:
    """
    Step 3 — Join scored leads to property details on OBJECTID.

    Left join from scores → leads ensures every scored property
    gets its details. Properties in leads that were not scored
    (filtered out by cleaner as non-residential) are excluded.
    The result is a single flat denormalized table — exactly
    what Power BI expects.
    """
    # Ensure OBJECTID is integer in both for a clean join
    scores["OBJECTID"] = scores["OBJECTID"].astype(int)
    leads["OBJECTID"]  = leads["OBJECTID"].astype(int)

    # Ensure scores has only the columns we need for the join
    score_cols = ["OBJECTID", "hot_lead_score", "rank"]
    available  = [c for c in score_cols if c in scores.columns]
    scores_slim = scores[available].copy()

    merged = scores_slim.merge(leads, on="OBJECTID", how="left")

    logger.info(
        f"Step 3 | Join scores+leads  | "
        f"{len(merged):,} rows | "
        f"matched={merged['OWN_NAME'].notna().sum():,}"
    )
    return merged


def _step4_temperature(df: pd.DataFrame) -> pd.DataFrame:
    """
    Step 4 — Assign temperature band based on score.

    This is a presentation-layer label for Power BI filtering
    and conditional formatting. It does NOT affect the model,
    the score, or the rank. The thresholds are:
        HOT  : score >= 0.8  (high-confidence motivated seller)
        WARM : score >= 0.5  (moderate confidence)
        COLD : score <  0.5  (low confidence)
    """
    def _assign(score):
        if score >= TEMP_HOT_THRESHOLD:
            return "HOT"
        elif score >= TEMP_WARM_THRESHOLD:
            return "WARM"
        return "COLD"

    df["temperature"] = df["hot_lead_score"].apply(_assign)

    hot  = (df["temperature"] == "HOT").sum()
    warm = (df["temperature"] == "WARM").sum()
    cold = (df["temperature"] == "COLD").sum()

    logger.info(
        f"Step 4 | Temperature bands  | "
        f"HOT={hot:,} | WARM={warm:,} | COLD={cold:,}"
    )
    return df


def _step5_format(df: pd.DataFrame) -> pd.DataFrame:
    """
    Step 5 — Reorder columns and cast data types for Power BI.

    Column order: rank → score → temperature → property details → OBJECTID.
    This order determines what Power BI shows in default table views.
    Integer columns are cast explicitly so Excel stores them as
    numbers, not text — Power BI auto-detects types from Excel.
    """
    # Only include columns that actually exist in the DataFrame
    final_cols = [c for c in EXPORT_COLUMNS if c in df.columns]
    export_df = df[final_cols].copy()

    # Cast integer columns
    int_cols = ["rank", "PHY_ZIPCD", "ACT_YR_BLT", "SALE_YR1",
                "NO_BULDNG", "OBJECTID"]
    for col in int_cols:
        if col in export_df.columns:
            export_df[col] = export_df[col].fillna(0).astype(int)

    # Sort by rank (should already be sorted, but enforce it)
    export_df = export_df.sort_values("rank").reset_index(drop=True)

    logger.info(
        f"Step 5 | Format columns     | "
        f"{len(export_df):,} rows × {len(final_cols)} columns"
    )
    return export_df


def _step6_save_to_db(df: pd.DataFrame):
    """
    Step 6 — Materialize the export as a DB table.

    Writes to export_leads table with if_exists='replace'.
    This table serves as:
        · Backup Power BI source (via ODBC if configured)
        · Ad-hoc SQL query target for analysis
        · Canonical record of what was exported this run
    """
    conn = sqlite3.connect(config.DB_NAME)
    try:
        df.to_sql("export_leads", conn, if_exists="replace", index=False)
        saved = pd.read_sql_query(
            "SELECT COUNT(*) AS total FROM export_leads", conn
        ).iloc[0]["total"]
        logger.info(
            f"Step 6 | Save to DB         | "
            f"export_leads → {saved:,} records"
        )
    except Exception as e:
        logger.error(f"Step 6 | DB save failed: {e}")
        raise
    finally:
        conn.close()


def _step7_write_excel(df: pd.DataFrame):
    """
    Step 7 — Write formatted Excel file for Power BI consumption.

    Creates an Excel workbook with:
        · A named Table ("ScoredLeads") — Power BI auto-detects
          named tables on import, no manual range selection needed
        · Frozen header row — always visible when scrolling
        · Auto-width columns — readable without manual resizing
        · TableStyleMedium9 — clean alternating-row style

    The file is written to the project folder. Power BI points
    at this file path. After each pipeline run, clicking Refresh
    in Power BI pulls the updated data automatically.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    excel_path = EXCEL_FILENAME

    # Write raw data first
    df.to_excel(excel_path, index=False, sheet_name=EXCEL_SHEET,
                engine="openpyxl")

    # Open and format
    wb = load_workbook(excel_path)
    ws = wb[EXCEL_SHEET]

    # Named table for Power BI auto-detection
    max_row = ws.max_row
    max_col = ws.max_column
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

    table = Table(displayName=TABLE_NAME, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Auto-width columns (header length + padding)
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        header_len = len(str(ws.cell(1, col_idx).value or ""))
        ws.column_dimensions[col_letter].width = max(header_len + 4, 12)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(excel_path)
    wb.close()

    file_size = os.path.getsize(excel_path) / (1024 * 1024)
    logger.info(
        f"Step 7 | Write Excel        | "
        f"{excel_path} → {file_size:.1f} MB | "
        f"Table: {TABLE_NAME}"
    )


# ============================================================
# MAIN NODE FUNCTION — called by main.py
# ============================================================

def run_export(state: dict) -> dict:
    """
    Main export node. Called by main.py orchestrator after predictor.

    Joins scored leads to property details from the leads table,
    adds temperature bands, writes a formatted Excel file for
    Power BI and an export_leads table in the database.

    Args:
        state : Shared pipeline State dict from build_state()

    Returns:
        Updated state with state['export_df'] populated.
        On failure, appends to state['errors'] and returns
        state with export_df = None.
    """
    logger.info("-" * 55)
    logger.info("NODE: run_export")
    logger.info("-" * 55)

    try:
        # ── Load and join ─────────────────────────────────────
        scores    = _step1_load_scores(state)
        leads     = _step2_load_lead_details(state)
        merged    = _step3_join(scores, leads)

        # ── Format for Power BI ───────────────────────────────
        merged    = _step4_temperature(merged)
        export_df = _step5_format(merged)

        # ── Write outputs ─────────────────────────────────────
        _step6_save_to_db(export_df)
        _step7_write_excel(export_df)

        # ── Write to state ────────────────────────────────────
        state["export_df"] = export_df

        hot  = (export_df["temperature"] == "HOT").sum()
        warm = (export_df["temperature"] == "WARM").sum()

        logger.info("-" * 55)
        logger.info(
            f"run_export complete | "
            f"{len(export_df):,} leads exported"
        )
        logger.info(
            f"Top lead: rank=1 | "
            f"score={export_df['hot_lead_score'].iloc[0]:.3f} | "
            f"{export_df['OWN_NAME'].iloc[0]}"
        )
        logger.info(
            f"Actionable leads: {hot + warm:,} "
            f"(HOT={hot:,} + WARM={warm:,})"
        )
        logger.info(
            f"Excel: {EXCEL_FILENAME} | "
            f"DB: export_leads table"
        )
        logger.info("-" * 55)

    except Exception as e:
        logger.error(f"run_export failed: {e}")
        state["errors"].append({
            "stage" : "export",
            "error" : str(e)
        })
        state["export_df"] = None

    return state
